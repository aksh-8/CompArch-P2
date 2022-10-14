import os
import glob
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm
statsfiles = glob.glob("D:/MS-CE-UTD/Sem-1/CE6304-Computer Architecture/Projects/Project2/*/*/*.txt")

L1d_stat = "system.cpu.dcache.overall_misses::total"
L1i_stat = "system.cpu.icache.overall_misses::total"
L2_stat = "system.l2.overall_misses::total"

for i in tqdm(range(len(statsfiles)//2)):

	t1 = os.path.dirname(statsfiles[i])
	t2 = os.path.basename(t1)
	t3 = t2.split("_")
	temp_dict = {}
	f = open(statsfiles[i],"r")
	l = f.readlines()
	for j in l[1:-2]+l[-1:]:
		k = j.split()
		key, values = k[0], k[1:]
		temp_dict[key] = values
		
	f.close()


	CPI = 1 + ((int(temp_dict[L1d_stat][0]) + int(temp_dict[L1i_stat][0])) * 6 + int(temp_dict[L2_stat][0]) * 50)/(500000000)

	workbook = load_workbook(filename = "All_Stats.xlsx")
	sheet_458 = workbook["458.sjeng"]
	sheet_458["C"+str(i+3)] = int(t3[1])
	sheet_458["D"+str(i+3)] = int(t3[2])
	sheet_458["E"+str(i+3)] = int(t3[4])
	sheet_458["F"+str(i+3)] = int(t3[6])
	sheet_458["G"+str(i+3)] = int(t3[8])
	sheet_458["H"+str(i+3)] = int(t3[10])
	sheet_458["I"+str(i+3)] = int(temp_dict[L1d_stat][0])
	sheet_458["J"+str(i+3)] = int(temp_dict[L1i_stat][0])
	sheet_458["K"+str(i+3)] = int(temp_dict[L2_stat][0])
	sheet_458["L"+str(i+3)] = CPI

	workbook.save(filename="All_Stats.xlsx")
	print([t2, L1d_stat, temp_dict[L1d_stat][0]])
	print([t2, L1i_stat, temp_dict[L1i_stat][0]])
	print([t2, L2_stat, temp_dict[L2_stat][0]])

for i in tqdm(range(len(statsfiles)//2, len(statsfiles))):

	t1 = os.path.dirname(statsfiles[i])
	t2 = os.path.basename(t1)
	t3 = t2.split("_")
	temp_dict = {}
	f = open(statsfiles[i],"r")
	l = f.readlines()
	for j in l[1:-2]+l[-1:]:
		k = j.split()
		key, values = k[0], k[1:]
		temp_dict[key] = values
		
	f.close()
	CPI = 1 + ((int(temp_dict[L1d_stat][0]) + int(temp_dict[L1i_stat][0])) * 6 + int(temp_dict[L2_stat][0]) * 50)/500000000

	workbook = load_workbook(filename = "All_Stats.xlsx")
	sheet_470 = workbook["470.lbm"]
	sheet_470["C"+str(i+3-144)] = int(t3[1])
	sheet_470["D"+str(i+3-144)] = int(t3[2])
	sheet_470["E"+str(i+3-144)] = int(t3[4])
	sheet_470["F"+str(i+3-144)] = int(t3[6])
	sheet_470["G"+str(i+3-144)] = int(t3[8])
	sheet_470["H"+str(i+3-144)] = int(t3[10])
	sheet_470["I"+str(i+3-144)] = int(temp_dict[L1d_stat][0])
	sheet_470["J"+str(i+3-144)] = int(temp_dict[L1i_stat][0])
	sheet_470["K"+str(i+3-144)] = int(temp_dict[L2_stat][0])
	sheet_470["L"+str(i+3-144)] = CPI

	workbook.save(filename="All_Stats.xlsx")
	print([t2, L1d_stat, temp_dict[L1d_stat][0]])
	print([t2, L1i_stat, temp_dict[L1i_stat][0]])
	print([t2, L2_stat, temp_dict[L2_stat][0]])
